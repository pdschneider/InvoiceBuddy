# Managers/history_manager.py
import logging, csv, os, shutil
from tkinter import messagebox
from openpyxl import load_workbook
from Utils.load_settings import load_history_path

headers = ["File Name", "Source Folder", "Destination Folder", "Type", "Moved", "Entered"]

def load_history(history_tree):
    """Loads history from the history.csv file, return default headers if file doesn't exist."""
    path = load_history_path()
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        # Create a new history file with headers if missing
        if not os.path.isfile(path):
            with open(path, "w", newline="", encoding="utf-8") as history:
                writer = csv.writer(history)
                writer.writerow(headers)
            logging.info(f"Created a new history file at {path}")
        # Write headers if file is empty
        if os.path.isfile(path) and os.path.getsize(path) == 0:
                    with open(path, "w", newline="", encoding="utf-8") as history:
                        writer = csv.writer(history)
                        writer.writerow(headers)
                    logging.info(f"Added headers to empty history file at {path}")
        # Loads data into the treeview
        history_tree.delete(*history_tree.get_children())
        with open(path, "r", newline="", encoding="utf-8") as file:
                reader = csv.reader(file)
                next(reader, None) #  Skips the header row
                for row in reader:
                    if len(row) == 6:
                        history_tree.insert("", "end", values=tuple(row))
                    else:
                            logging.warning(f"Skipping invalid row in history.csv: {row}")
        logging.info(f"Loaded {len(history_tree.get_children())} entries from {path}")
    except Exception as e:
            logging.error(f"Could not load history file due to: {e}")

def add_update_history(filename, src_folder, dst_folder=None, file_type=None, moved=None, entered=None):
    """Adds or updates a history entry. Only changes fields that are provided (not None)."""
    path = load_history_path()
    rows = []

    # Read existing CSV
    if os.path.isfile(path) and path.lower().endswith("csv"):
        with open(path, "r", newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

    # Ensure headers exist
    if not rows or rows[0] != headers:
        rows = [headers] + (rows if rows else [])

    # Find existing row
    row_index = None
    for i in range(1, len(rows)):
        if len(rows[i]) > 0 and rows[i][0] == filename:
            row_index = i
            break

    if row_index is not None:
        # Update existing row — preserve old values unless new one provided
        current = rows[row_index]
        # Pad short rows
        while len(current) < 6:
            current.append("")

        new_row = [
            filename,                                      # column 0
            src_folder if src_folder is not None else current[1],
            dst_folder if dst_folder is not None else current[2],
            file_type if file_type is not None else current[3],
            moved if moved is not None else current[4],
            entered if entered is not None else current[5]]
        rows[row_index] = new_row
        logging.debug(f"Updated history entry for {filename}")
    else:
        # New entry — use defaults where nothing provided
        new_row = [
            filename,
            src_folder or "",           # or raise error if required?
            dst_folder or "N/A",
            file_type or "",            # or raise error if required?
            moved or "No",
            entered or "No"]
        rows.append(new_row)
        logging.debug(f"Added new history entry for {filename}")

    # Write back
    try:
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerows(rows)
        logging.info(f"Successfully updated history.csv for {filename}")
    except Exception as e:
        logging.error(f"Failed to update history.csv for {filename}: {e}")

def revert_moves(history_tree):
    """Reverts files back to their previous directory location."""
    selected_ids = history_tree.selection()
    if not selected_ids:
         logging.warning(f"Attempted to revert moves with no selection in history treeview.")
         return
    path = load_history_path()
    rows = []
    errors = []

    # Load CSV rows
    with open(path, "r", newline="", encoding="utf-8") as file:
         reader = csv.reader(file)
         rows = list(reader)

    selected_filenames = [history_tree.item(item_id)["values"][0] for item_id in selected_ids]

    for filename in selected_filenames:
        for row in rows[1:]: #  Skips the header
            if row[0] == filename and row[4] == "Yes":
                src_folder = row[1]
                dst_folder = row[2]
                dst_path = os.path.join(dst_folder, filename)
                src_path = os.path.join(src_folder, filename)
                if os.path.exists(dst_path):
                    try:
                            shutil.move(dst_path, src_path)
                            row[4] = "No"
                            row[2] = "N/A"
                    except Exception as e:
                            errors.append(f"Failed to revert {filename} due to: {e}")
                else:
                    errors.append(f"File not found at {dst_path}")
                break
    # Writes updated rows back to CSV
    with open(path, "w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerows(rows)

    # Reload Treeview
    load_history(history_tree)

    if errors:
        messagebox.showerror("Errors", "\n".join(errors))
    else:
        messagebox.showinfo("Success", "Selected file moves reverted.")

def remove_spreadsheet_entries(history_tree, globals):
    """Removes selected entries from the spreadsheet."""
    selected_ids = history_tree.selection()
    if not selected_ids:
        logging.warning(f"Attmpted to remove spreadsheet entries without selecting anything.")
        return

    path = load_history_path()
    rows = []
    with open(path, "r", newline="", encoding="utf-8") as file:
        reader = csv.reader(file)
        rows = list(reader)

    selected_filenames = [history_tree.item(item_id)["values"][0] for item_id in selected_ids]

    try:
        wb = load_workbook(globals.workbook)
        sheet_map = {
            "Invoices": wb[globals.sheet_invoices],
            "Credit Cards": wb[globals.sheet_CreditCards],
            "Purchase Orders": wb[globals.sheet_PurchaseOrders]}
        removed = False

        for filename in selected_filenames:
            for row_idx, row in enumerate(rows[1:], start=1):
                if row[0] == filename and row[5] == "Yes":
                    file_type = row[3]
                    if file_type not in sheet_map:
                        continue
                    sheet = sheet_map[file_type]
                    base_filename = os.path.splittext(filename)[0]
                    portions = base_filename.split()
                    start_row = 3 if file_type in ["Invoices", "Purchase Orders"] else 8
                    for sheet_row in range(start_row, sheet.max_row + 1):
                        row_values = [str(sheet.cell(row=sheet_row, column=col).value or "").strip().lower() for col in range(1, len(portions) + 1)]
                        if row_values == [p.lower().strip() for p in portions]:
                            for col in range(1, len(portions) + 1):
                                sheet.cell(row=sheet_row, column=col).value = None
                            removed = True
                            rows[row_idx][5] = "No"
                            break

        if os.access(globals.workbook, os.W_OK):
            wb.save(globals.workbook)
            # Write updated rows back to history.csv
            with open(path, "w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                writer.writerows(rows)
            # Reload Treeview
            load_history(history_tree)
            if removed:
                messagebox.showinfo("Success", "Spreadsheet entries removed!")
            else:
                messagebox.showwarning("Warning", "No matching spreadsheet entries found to remove.")
        else:
            raise PermissionError(f"Permission denied to write to {globals.workbook}")
    except Exception as e:
        logging.error(f"Failed to remove spreadsheet entries due to: {e}")
        messagebox.showerror("Error", f"Failed to remove spreadsheet entries due to: {e}")

def remove_from_history(history_tree):
    """Removes reverted entries from the history treeview."""
    selected_ids = history_tree.selection()
    if not selected_ids:
        logging.warning(f"No items selected to remove from the history log.")
        return

    path = load_history_path()
    rows = []
    with open(path, "r", newline="", encoding="utf-8") as file:
        reader = csv.reader(file)
        rows = list(reader)

    selected_filenames = [history_tree.item(item_id)["values"][0] for item_id in selected_ids]
    updated_rows = [rows[0]] + [row for row in rows[1:] if row[0] not in selected_filenames]

    with open(path, "w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerows(updated_rows)

    load_history(history_tree)
    logging.info(f"Selected entries removed from the history log.")
