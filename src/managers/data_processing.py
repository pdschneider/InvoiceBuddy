# src/managers/data_processing.py
import os
import logging
import msoffcrypto
from io import BytesIO
from tkinter import simpledialog
from openpyxl import load_workbook
from src.managers.history_manager import load_history, add_update_history
from src.utils.toast import show_toast


def parse_sheet_qt(globals, sheet_name, file_list=None):
    """
    Qt-mode parser: writes filename data to a user-defined sheet.
    Reads sheet config (workbook, starting row/col, column order) from sheet_data.
    """

    # Exit early if no files are selected
    if not file_list:
        show_toast(globals, "Please select one or more files to enter.")
        return

    # Get sheet config (name, color, workbook path, first row/column, column order)
    sheet_config = None
    for sheet in globals.sheet_data.get("sheets", []):
        if sheet.get("name") == sheet_name:
            sheet_config = sheet
            logging.debug(f"Sheet Config: {sheet_config}")
            break

    # Exit if sheet name is not found in sheet config
    if not sheet_config:
        logging.error(f"Sheet '{sheet_name}' not found in sheet definitions.")
        return

    workbook_path = sheet_config.get("workbook", "")
    if not os.path.isfile(workbook_path):
        show_toast(globals, f"No valid workbook for sheet '{sheet_name}'.", _type="error")
        return

    # Use correct workbook
    if not globals.legacy_mode:
        globals.workbook = workbook_path

    # Extract workbook to memory object and return if empty
    wb = encryption_handler(globals)
    if not wb:
        return

    # Ensure chosen sheet name is in workbook sheet names
    wb_sheets = wb.sheetnames
    logging.debug(f"Available Sheets: {wb_sheets}")
    if sheet_name not in wb_sheets:
        logging.warning(f"{sheet_name} not found in workbook.")
        show_toast(globals, message=f"{sheet_name} not in workbook")
        return

    try:
        base_names = [os.path.basename(f) for f in file_list]
        sheet = wb[sheet_config.get("sheet_name", sheet_name)]
        logging.debug(f"Sheet Name: {sheet}")

        starting_row = sheet_config.get("first_row", 3)
        starting_column = sheet_config.get("first_column", 1)
        column_order = sheet_config.get("column_order", ["Company", "Date", "Invoice #", ""])

        columns_to_check = 5

        while not all(
            sheet.cell(row=starting_row, column=col).value is None
            for col in range(starting_column, starting_column + columns_to_check)):
            starting_row += 1

        for full_file_name in base_names:
            while not all(
                sheet.cell(row=starting_row, column=col).value is None
                for col in range(starting_column, starting_column + columns_to_check)):
                starting_row += 1

            base_name = os.path.splitext(full_file_name)[0]
            portions = base_name.split()

            for j, portion in enumerate(portions, start=starting_column):
                sheet.cell(row=starting_row, column=j, value=portion.strip())

            starting_row += 1

        if os.access(workbook_path, os.W_OK):
            wb.save(workbook_path)
            logging.info(f"Saved {len(base_names)} entries to '{sheet_name}' in {workbook_path}")
        else:
            logging.warning(f"Permission denied to write to {workbook_path}")

    except Exception as e:
        logging.error(f"Qt sheet parsing error for '{sheet_name}': {e}")
        show_toast(globals, message=f"Error: '{sheet_name}': {e}", _type="error")


def paths_check(globals):
    """Ensures that the workbook is writable and inbox path is valid."""
    try:
        # Check if inbox path is valid
        if os.path.isdir(globals.inbox):
            pass
        elif os.path.isdir(globals.inbox_dir_var.get().strip()):
            globals.inbox = globals.inbox_dir_var.get().strip()
        else:
            logging.warning(f"Inbox path not valid - Skipping entering data.")
            return False

        # Check if workbook path is valid
        if os.path.isfile(globals.workbook):
            pass
        elif os.path.isfile(globals.workbook_var.get().strip()):
            globals.workbook = globals.workbook_var.get().strip()
        else:
            logging.warning(f"Workbook path not valid - Skipping entering data.")
            return False

        # Ensure workbook has writable filesystem permissions
        # This only checks for general filesystem permissions,
        # NOT file locks such as when another user has the spreadhseet open.
        if os.access(globals.workbook, os.W_OK):
            pass
        else:
            logging.error(f"Permission denied to write to workbook via filesystem permissions.")
            return False
        
        # Return True only if checks pass
        return True

    except Exception as e:
        logging.error(f"Unable to check paths due to: {e}")
        return False


def encryption_handler(globals):
    """
    Returns a wb object and handles encrypted workbooks.
    
        is_encrypted:   True if workbook is password protected, else False
        wb:             Workbook object in memory
    """
    is_encrypted = False

    try:
        # Check to see if workbook is encrypted
        with open(globals.workbook, "rb") as f:
            workbook_file = msoffcrypto.OfficeFile(f)
            is_encrypted = workbook_file.is_encrypted()

        # Handle password entry if workbook is encrypted
        if is_encrypted:
            password = simpledialog.askstring(
                "Password Required",
                "Enter the password for the encrypted workbook:",
                show='*')
            if password:
                decrypted = BytesIO()
                try:
                    with open(globals.workbook, 'rb') as f:
                        file = msoffcrypto.OfficeFile(f)
                        file.load_key(password=password)
                        file.decrypt(decrypted)
                except:
                    logging.error(f"Password failed.")
                    return False

                # Return if password entry failed
                if is_encrypted:
                    logging.error(f"Decryption failed.")
                    show_toast(globals, message="Decryption failed - wrong password?")
                    return False

                # Return decrypted workbook
                decrypted.seek(0)
                wb = load_workbook(decrypted)
                return wb

            else: # If password is not entered
                logging.info(f"Exited password entry.")
                return False
        else: # If not encrypted
            wb = load_workbook(globals.workbook)
            return wb
    except Exception as e:
        logging.error(f"Unable to generate workbook object due to: {e}")
        return False


def parse_invoices(globals, history_tree, file_list=None):
    """
    Writes filename data to the Invoices sheet of the workbook,
    respecting a configurable starting column.

        globals:        Global variables
        history_tree:   Tkinter treeview
        file_list:      list of filepath strings from the inbox view
                        ex: ['/home/phillip/Phillip Inbox/03-04-26 109215.pdf']

        base_names:     A list of base names for each entry in file_list
                        ex: ['03-04-26 109215.pdf']
    """
    # Return if inbox or workbook paths are not valid
    if not paths_check(globals):
        return

    # Return if file list is empty
    if not file_list:
        show_toast(globals, "Please select one or more files to enter.")
        return

    # Generate wb object, return if empty
    wb = encryption_handler(globals)
    if not wb:
        return

    try:
        base_names = [os.path.basename(f) for f in file_list]

        # Generate sheet object
        sheet = wb[globals.sheet_invoices]

        # Set starting row
        if globals.invoice_starting_row and isinstance(globals.invoice_starting_row, int):
            current_row = globals.invoice_starting_row
        else:
            current_row = 1
            logging.error(
                f"Could not read invoice starting row. Defaulting to 1.")

        # Set starting column
        if globals.invoice_starting_column and isinstance(globals.invoice_starting_column, int):
            starting_column = globals.invoice_starting_column
        else:
            starting_column = 1
            logging.error(
                "Could not read invoice starting column. Defaulting to 1.")

        columns_to_check = 5  # If any of these columns are not empty, skip row

        # Find the first fully empty row starting from current_row
        while not all(
            sheet.cell(row=current_row, column=col).value is None
            for col in range(starting_column, starting_column + columns_to_check)):
            current_row += 1
        logging.info(
            f"First available row: {current_row} (starting at column {starting_column})")

        # Process each file
        for full_file_name in base_names:
            # Double-check row is empty before writing
            while not all(sheet.cell(row=current_row, column=col).value is None for col in range(starting_column, starting_column + columns_to_check)):
                current_row += 1

            # Create list based on filename, separated by spaces
            base_name = os.path.splitext(full_file_name)[0]
            portions = base_name.split()

            # Write each portion starting at the defined column (in memory)
            for j, portion in enumerate(portions, start=starting_column):
                sheet.cell(row=current_row, column=j, value=portion.strip())

            current_row += 1  # Move to next row for next file

        # Update history for all processed files
        for full_file_name in base_names:
            add_update_history(
                filename=full_file_name,
                src_folder=globals.inbox,
                file_type="Invoices",
                entered=globals.user)

        # Save workbook
        if os.access(globals.workbook, os.W_OK):
            wb.save(globals.workbook)

            # Refresh history treeview
            load_history(history_tree)
        else:
            logging.warning(f"Permission denied to write to {globals.workbook}")
            raise PermissionError(
                f"Permission denied to write to {globals.workbook}")

    except Exception as e:
        logging.error(f"{globals.invoice_sheet_label} processing error: {e}")


def parse_credit_cards(globals, history_tree, file_list=None):
    """
    Writes filename data to the Credit Cards sheet of the workbook,
    respecting a configurable starting column.

        globals:        Global variables
        history_tree:   Tkinter treeview
        file_list:      list of filepaths from the inbox view
    """
    # Return if inbox or workbook paths are not valid
    if not paths_check(globals):
        return

    # Return if file list is empty
    if not file_list:
        show_toast(globals, "Please select one or more files to enter.")
        return

    # Generate wb object, return if empty
    wb = encryption_handler(globals)
    if not wb:
        return

    try:
        base_names = [os.path.basename(f) for f in file_list]

        # Generate sheet object
        sheet = wb[globals.sheet_CreditCards]

        # Starting row (fallback to 3)
        if globals.card_starting_row and isinstance(globals.card_starting_row, int):
            current_row = globals.card_starting_row
        else:
            current_row = 3


        # Starting column (fallback to 1)
        if globals.card_starting_column and isinstance(globals.card_starting_column, int):
            starting_column = globals.card_starting_column
        else:
            starting_column = 1

        columns_to_check = 3 # Ensure at least the first 3 columns are empty

        # Find the first fully empty row starting from current_row
        while not all(
            sheet.cell(row=current_row, column=col).value is None
            for col in range(starting_column, starting_column + columns_to_check)):
            current_row += 1
        logging.info(
            f"First available row: {current_row} (starting at column {starting_column})")

        # Process each file
        for full_file_name in base_names:
            # Double-check row is empty before writing
            while not all(
                sheet.cell(row=current_row, column=col).value is None
                for col in range(starting_column, starting_column + columns_to_check)):
                current_row += 1

            # Generate a list split by spaces in filename
            base_name = os.path.splitext(full_file_name)[0]
            portions = base_name.split()

            # Write each portion starting at the defined column
            for j, portion in enumerate(portions, start=starting_column):
                sheet.cell(row=current_row, column=j, value=portion.strip())

            current_row += 1  # Move to next row for next file

        # Update history for all processed files
        for full_file_name in base_names:
            add_update_history(
                filename=full_file_name,
                src_folder=globals.inbox,
                file_type="Credit Cards",
                entered=globals.user)

        # Save workbook
        if os.access(globals.workbook, os.W_OK):
            wb.save(globals.workbook)
            logging.info(f"Saved workbook to {globals.workbook}")

            # Refresh history treeview
            load_history(history_tree)
        else:
            logging.warning("Permission denied to write to workbook.")
            raise PermissionError(
                f"Permission denied to write to {globals.workbook}")

    except Exception as e:
        logging.error(f"{globals.card_sheet_label} processing error: {e}")
