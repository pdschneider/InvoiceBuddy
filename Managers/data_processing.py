#Managers/data_processing.py
import os, logging, msoffcrypto
from io import BytesIO
from tkinter import messagebox, simpledialog
from openpyxl import load_workbook
from Managers.history_manager import load_history, add_update_history

def parse_invoices(globals, history_tree, file_list=None):
    """Writes filename data to the Invoices sheet of the workbook, respecting a configurable starting column."""
    try:
        if os.path.isdir(globals.sources["inbox"]):
            invoice_dir = globals.sources["inbox"]
        elif os.path.isdir(globals.inbox_dir_var.get().strip()):
            invoice_dir = globals.inbox_dir_var.get().strip()
        else:
            logging.warning(f"No valid inbox path. Skipping entering data.")
            return
        if os.path.isfile(globals.workbook):
            workbook_file_path = globals.workbook
        elif os.path.isfile(globals.workbook_var.get().strip()):
            workbook_file_path = globals.workbook_var.get().strip()
        else:
            return
        logging.debug(f"Checking directory: {invoice_dir}")

        if file_list is None:
            all_files = os.listdir(invoice_dir)
            logging.debug(f"Directory contents: {all_files}")
            full_file_names = [f for f in all_files if os.path.isfile(os.path.join(invoice_dir, f))]
            logging.debug(f"Found files: {full_file_names}")
        else:
            full_file_names = [os.path.basename(f) for f in file_list]

        if not full_file_names:
            messagebox.showwarning("Warning", f"No files found in {invoice_dir}")
            logging.warning(f"No files found in {invoice_dir}")
            return

        # Load workbook (with password support for encrypted files)
        try:
            wb = load_workbook(workbook_file_path)
        except Exception as load_error:
            if "invalid" in str(load_error).lower() or "zip" in str(load_error).lower():
                password = simpledialog.askstring("Password Required", "Enter the password for the encrypted workbook:", show='*')
                if password:
                    decrypted = BytesIO()
                    with open(workbook_file_path, 'rb') as f:
                        file = msoffcrypto.OfficeFile(f)
                        file.load_key(password=password)
                        file.decrypt(decrypted)
                    decrypted.seek(0)
                    wb = load_workbook(decrypted)
                else:
                    raise
            else:
                raise

        sheet = wb[globals.sheet_invoices]
        logging.info(f"Loaded sheet '{globals.sheet_invoices}' with max row: {sheet.max_row}")

        try:
            current_row = globals.invoice_starting_row
        except Exception:
            current_row = 4
            logging.error("Could not read invoice starting row. Defaulting to 3.")

        if globals.invoice_starting_column:
            starting_column = globals.invoice_starting_column
        else:
            starting_column = 1
        columns_to_check = 5

        # Find the first fully empty row starting from current_row
        while not all(
            sheet.cell(row=current_row, column=col).value is None
            for col in range(starting_column, starting_column + columns_to_check)
        ):
            current_row += 1
        logging.info(f"First available row: {current_row} (starting at column {starting_column})")

        # Process each file
        for full_file_name in full_file_names:
            # Double-check row is empty before writing
            while not all(
                sheet.cell(row=current_row, column=col).value is None
                for col in range(starting_column, starting_column + columns_to_check)):
                current_row += 1

            base_name = os.path.splitext(full_file_name)[0]
            portions = base_name.split()
            logging.debug(f"Writing '{full_file_name}' → row {current_row}, columns {starting_column}+ : {portions}")

            # Write each portion starting at the defined column
            for j, portion in enumerate(portions, start=starting_column):
                sheet.cell(row=current_row, column=j, value=portion.strip())

            current_row += 1  # Move to next row for next file

        # Update history for all processed files
        for full_file_name in full_file_names:
            add_update_history(
                filename=full_file_name,
                src_folder=invoice_dir,
                file_type="Invoices",
                entered=globals.user
            )

        # Save workbook
        if os.access(workbook_file_path, os.W_OK):
            wb.save(workbook_file_path)
            logging.info(f"Saved workbook to {workbook_file_path}")
            messagebox.showinfo("Success", "Invoice data was added to the spreadsheet.")

            # Refresh history treeview
            load_history(history_tree)
        else:
            raise PermissionError(f"Permission denied to write to {workbook_file_path}")

    except Exception as e:
        logging.error(f"Invoice processing error: {e}")
        messagebox.showerror("Error", f"An error occurred while processing invoices:\n{e}")

def parse_credit_cards(globals, history_tree, file_list=None):
    """Writes filename data to the Credit Cards sheet of the workbook, respecting a configurable starting column."""
    try:
        if os.path.isdir(globals.sources["inbox"]):
            credit_dir = globals.sources["inbox"]
        elif os.path.isdir(globals.inbox_dir_var.get().strip()):
            credit_dir = globals.inbox_dir_var.get().strip()
        else:
            logging.warning(f"No valid inbox path. Skipping entering data.")
            return
        if os.path.isfile(globals.workbook):
            workbook_file_path = globals.workbook
        elif os.path.isfile(globals.workbook_var.get().strip()):
            workbook_file_path = globals.workbook_var.get().strip()
        else:
            return
        logging.debug(f"Checking directory: {credit_dir}")

        if file_list is None:
            all_files = os.listdir(credit_dir)
            logging.debug(f"Directory contents: {all_files}")
            full_file_names = [f for f in all_files if os.path.isfile(os.path.join(credit_dir, f))]
            logging.debug(f"Found files: {full_file_names}")
        else:
            full_file_names = [os.path.basename(f) for f in file_list]

        if not full_file_names:
            messagebox.showwarning("Warning", f"No files found in {credit_dir}")
            logging.warning(f"No files found in {credit_dir}")
            return

        # Load workbook (with password support for encrypted files)
        try:
            wb = load_workbook(workbook_file_path)
        except Exception as load_error:
            if "invalid" in str(load_error).lower() or "zip" in str(load_error).lower():
                password = simpledialog.askstring("Password Required", "Enter the password for the encrypted workbook:", show='*')
                if password:
                    decrypted = BytesIO()
                    with open(workbook_file_path, 'rb') as f:
                        file = msoffcrypto.OfficeFile(f)
                        file.load_key(password=password)
                        file.decrypt(decrypted)
                    decrypted.seek(0)
                    wb = load_workbook(decrypted)
                else:
                    raise
            else:
                raise

        sheet = wb[globals.sheet_CreditCards]
        logging.info(f"Loaded sheet '{globals.sheet_CreditCards}' with max row: {sheet.max_row}")

        # Starting row from settings (fallback to 3)
        try:
            current_row = globals.card_starting_row
        except Exception:
            current_row = 3
            logging.error("Could not read card starting row. Defaulting to 3.")

        if globals.card_starting_column:
            starting_column = globals.card_starting_column
        else:
            starting_column = 1
        columns_to_check = 3

        # Find the first fully empty row starting from current_row
        while not all(
            sheet.cell(row=current_row, column=col).value is None
            for col in range(starting_column, starting_column + columns_to_check)
        ):
            current_row += 1
        logging.info(f"First available row: {current_row} (starting at column {starting_column})")

        # Process each file
        for full_file_name in full_file_names:
            # Double-check row is empty before writing
            while not all(
                sheet.cell(row=current_row, column=col).value is None
                for col in range(starting_column, starting_column + columns_to_check)
            ):
                current_row += 1

            base_name = os.path.splitext(full_file_name)[0]
            portions = base_name.split()
            logging.debug(f"Writing '{full_file_name}' → row {current_row}, columns {starting_column}+ : {portions}")

            # Write each portion starting at the defined column
            for j, portion in enumerate(portions, start=starting_column):
                sheet.cell(row=current_row, column=j, value=portion.strip())

            current_row += 1  # Move to next row for next file

        # Update history for all processed files
        for full_file_name in full_file_names:
            add_update_history(
                filename=full_file_name,
                src_folder=credit_dir,
                file_type="Credit Cards",
                entered=globals.user
            )

        # Save workbook
        if os.access(workbook_file_path, os.W_OK):
            wb.save(workbook_file_path)
            logging.info(f"Saved workbook to {workbook_file_path}")
            messagebox.showinfo("Success", "Credit card data was added to the spreadsheet.")

            # Refresh history treeview
            load_history(history_tree)
        else:
            raise PermissionError(f"Permission denied to write to {workbook_file_path}")

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while processing credit cards:\n{e}")
        logging.error(f"Credit card processing error: {e}")
